"""
remplir_excel.py
================
Remplit intelligemment le fichier SUIVI_HEBDO_JLD_2026.xlsx via Claude API.
Chaque mission est routée vers sa zone de colonnes dans l'onglet hebdomadaire :
  - RT Compteur (F-P)  : RT_Compteur_Module, RT_CPT_Arras, RT_CPT_SEPIG, RT_CPT_Suez, RT_CPT
  - Contrôle AC/ANC (U-AA) : Controle_AC, Controle_ANC

Utilisation : python remplir_excel.py
"""

import json
import sys
import logging
import anthropic
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
from datetime import date
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from firebase_client import charger_techniciens, charger_saisies_semaine

BASE_DIR   = Path(__file__).parent
EXCEL_SRC  = BASE_DIR.parent / "SUIVI HEBDO JLD 2026 (1).xlsx"
EXPORT_DIR = BASE_DIR / "exports"

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

# ─── MAPPING COLONNES PAR ZONE ────────────────────────────────────────────────
# Zone RT Compteur (F-P) — missions RT_*
COL_RT = {
    "cpt_dn15_20":      "F",   # RTCT DN 15-20
    "cpt_dn_sup20":     "G",   # RTCT DN 30-40
    "mo_heures":        "H",   # MO
    "modules_poses":    "I",   # TOTAL MODULES changés
    "modules_relances": "J",   # Modules relancés
    "rac":              "K",   # RAC
    "clapets_fact":     "L",   # TP
    "sr_laiton":        "M",   # SR
    "reducteurs":       "N",   # RP
    "clients_absents":  "O",   # INF / AVISÉS
    "depl_injustifies": "P",   # DI
}

# Zone Contrôle AC / ANC (U-AA) — missions Controle_*
COL_CTRL = {
    "ctrl_vente_inf10":      "U",   # CTRL AC (total <10pts)
    "ctrl_vente_sup10":      "V",   # AC > 10 pts
    "ctrl_contrat_inf10":    "W",   # CTRL ANC (<10pts)
    "ctrl_contrat_sup10":    "X",   # ANC > 10 pts (ENQUETES)
    "rdv_eae":               "Y",   # RAPPORT
    "clients_absents":       "Z",   # INF / AVISÉS
    "depl_injustifies":      "AA",  # DI
}

# Routing mission → zone de colonnes
MISSIONS_RT   = {"RT_Compteur_Module", "RT_CPT_Arras", "RT_CPT_SEPIG", "RT_CPT_Suez", "RT_CPT"}
MISSIONS_CTRL = {"Controle_AC", "Controle_ANC"}

def col_mapping_pour_mission(mission):
    if mission in MISSIONS_RT:
        return COL_RT
    if mission in MISSIONS_CTRL:
        return COL_CTRL
    return {}  # Mission non gérée (ex: Releve_CPT → onglet Sem_X, non implémenté ici)


# ─── UTILITAIRES ──────────────────────────────────────────────────────────────
def get_semaine_courante():
    return date.today().isocalendar()[1]

def get_annee_courante():
    return date.today().year

def get_semaine_key():
    return f"{get_annee_courante()}-S{get_semaine_courante()}"


def calculer_totaux(techniciens, saisies):
    """Calcule les totaux hebdomadaires par technicien."""
    champs_num = [
        "cpt_dn15_20", "cpt_dn_sup20", "mo_heures", "modules_poses",
        "modules_relances", "rac", "clapets_fact", "sr_laiton", "reducteurs",
        "clients_absents", "depl_injustifies", "total_heures",
    ]
    totaux = {}
    for tech in techniciens:
        uid = tech.get("uid") or tech.get("id")
        mes_saisies = [s for s in saisies if s.get("tech_id") == uid]
        total = {"uid": uid, "nom": tech.get("nom"), "prenom": tech.get("prenom"),
                 "mission": tech.get("mission"), "nb_jours": len(mes_saisies)}
        for champ in champs_num:
            total[champ] = sum(s.get(champ, 0) or 0 for s in mes_saisies)
        totaux[uid] = total
    return totaux


def lire_lignes_excel(ws):
    """Extrait les lignes contenant des noms de techniciens (col E non vide, col D = 1)."""
    lignes = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        col_d = row[3].value  # colonne D
        col_e = row[4].value  # colonne E
        if col_d == 1 and col_e and str(col_e).strip():
            lignes.append({"row": row[0].row, "nom_excel": str(col_e).strip()})
    return lignes


def matcher_techniciens_claude(lignes_excel, totaux):
    """
    Utilise Claude API pour matcher les noms Firebase aux lignes Excel.
    Retourne: {uid: row_number}
    """
    missions_gerees = MISSIONS_RT | MISSIONS_CTRL
    firebase_techs = [
        {"uid": uid, "nom": t["nom"], "prenom": t["prenom"], "mission": t["mission"]}
        for uid, t in totaux.items() if t["mission"] in missions_gerees
    ]

    if not firebase_techs:
        log.info("Aucun technicien avec une mission gérée (RT_* ou Controle_*).")
        return {}

    prompt = f"""Tu es un assistant qui mappe des données.

Voici les techniciens dans Firebase (RT Compteur) :
{json.dumps(firebase_techs, ensure_ascii=False, indent=2)}

Voici les lignes de techniciens dans le fichier Excel (numéro de ligne et nom affiché) :
{json.dumps(lignes_excel, ensure_ascii=False, indent=2)}

Pour chaque technicien Firebase, trouve la ligne Excel correspondante en faisant correspondre les noms
(ignore les accents, espaces superflus, suffixes comme 'apl', 'sms', 'mail').

Retourne UNIQUEMENT un objet JSON valide (pas de texte autour) de cette forme :
{{
  "uid_du_technicien": numero_de_ligne,
  ...
}}

Si un technicien Firebase ne correspond à aucune ligne Excel, ne l'inclus pas dans le résultat.
"""

    client = anthropic.Anthropic()
    response = client.messages.create(
        model="claude-opus-4-8",
        max_tokens=512,
        messages=[{"role": "user", "content": prompt}]
    )

    raw = response.content[0].text.strip()
    # Extraire le JSON si entouré de markdown
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    mapping = json.loads(raw)
    log.info(f"Claude a matché {len(mapping)} techniciens sur {len(firebase_techs)}")
    return {k: int(v) for k, v in mapping.items()}


# ─── LOGIQUE PRINCIPALE ───────────────────────────────────────────────────────
def remplir_excel(semaine=None, annee=None):
    semaine = semaine or get_semaine_courante()
    annee   = annee   or get_annee_courante()
    nom_onglet = str(semaine)

    log.info(f"=== Remplissage Excel S{semaine} {annee} ===")

    # 1. Charger données Firebase
    log.info("Chargement des données Firebase...")
    techniciens = charger_techniciens()
    saisies     = charger_saisies_semaine()
    if not saisies:
        log.warning("Aucune saisie pour la semaine courante — Excel non modifié.")
        return None

    totaux = calculer_totaux(techniciens, saisies)
    nb_saisis = sum(1 for t in totaux.values() if t["nb_jours"] > 0)
    log.info(f"{len(totaux)} techniciens, {nb_saisis} avec au moins une saisie.")

    # 2. Ouvrir le fichier Excel
    log.info(f"Ouverture de : {EXCEL_SRC}")
    wb = openpyxl.load_workbook(str(EXCEL_SRC))

    # 3. Trouver ou créer l'onglet de la semaine
    if nom_onglet in wb.sheetnames:
        ws = wb[nom_onglet]
        log.info(f"Onglet '{nom_onglet}' trouvé.")
    else:
        # Copier depuis le dernier onglet numérique disponible
        onglets_num = sorted([int(s) for s in wb.sheetnames if s.isdigit()])
        if not onglets_num:
            log.error("Aucun onglet numérique trouvé dans le fichier Excel.")
            return None
        dernier = str(onglets_num[-1])
        ws = wb.copy_worksheet(wb[dernier])
        ws.title = nom_onglet
        # Mettre à jour le numéro de semaine en cellule J1
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=False):
            for cell in row:
                if cell.value == onglets_num[-1]:
                    cell.value = semaine
        # Vider les cellules de données (F à P, lignes techniciens)
        lignes_techs = lire_lignes_excel(ws)
        toutes_cols = list(COL_RT.values()) + list(COL_CTRL.values())
        for lt in lignes_techs:
            for col_lettre in toutes_cols:
                ws[f"{col_lettre}{lt['row']}"] = None
        log.info(f"Onglet '{nom_onglet}' créé par copie de '{dernier}' et nettoyé.")

    # 4. Lire les lignes de techniciens dans l'onglet
    lignes_excel = lire_lignes_excel(ws)
    log.info(f"{len(lignes_excel)} lignes techniciens trouvées dans l'onglet.")

    # 5. Claude mappe les UIDs Firebase aux lignes Excel
    mapping_uid_row = matcher_techniciens_claude(lignes_excel, totaux)
    if not mapping_uid_row:
        log.warning("Aucun technicien matché par Claude.")
        return None

    # 6. Écrire les valeurs dans les cellules selon la zone de la mission
    nb_remplis = 0
    for uid, row_num in mapping_uid_row.items():
        total = totaux.get(uid)
        if not total:
            continue
        nom_complet = f"{total['prenom']} {total['nom']}"
        mission = total.get("mission", "")
        col_mapping = col_mapping_pour_mission(mission)
        zone = "RT F-P" if mission in MISSIONS_RT else "CTRL U-AA" if mission in MISSIONS_CTRL else "?"

        if not col_mapping:
            log.warning(f"  ? {nom_complet} — mission '{mission}' sans mapping de colonnes")
            continue

        cellules_ecrites = []
        for champ_firebase, col_lettre in col_mapping.items():
            valeur = total.get(champ_firebase, 0)
            if valeur:
                cell = ws[f"{col_lettre}{row_num}"]
                cell.value = valeur
                # Force la couleur noire — la copie d'onglet hérite parfois d'un texte blanc
                cell.font = Font(color="000000")
                cellules_ecrites.append(f"{col_lettre}{row_num}={valeur}")

        if cellules_ecrites:
            log.info(f"  ✓ {nom_complet} [{zone}] (ligne {row_num}, {total['nb_jours']}j) → {', '.join(cellules_ecrites)}")
        else:
            log.info(f"  – {nom_complet} [{zone}] (ligne {row_num}) → aucune valeur (0 jours saisis)")
        nb_remplis += 1

    log.info(f"{nb_remplis} techniciens traités dans l'Excel.")

    # 7. Sauvegarder (gère le cas où le fichier est déjà ouvert)
    EXPORT_DIR.mkdir(exist_ok=True)
    fichier_sortie = EXPORT_DIR / f"SUIVI_S{semaine}_{annee}.xlsx"
    try:
        wb.save(str(fichier_sortie))
        log.info(f"Fichier sauvegardé : {fichier_sortie}")
    except PermissionError:
        from datetime import datetime
        alt = EXPORT_DIR / f"SUIVI_S{semaine}_{annee}_{datetime.now().strftime('%H%M%S')}.xlsx"
        wb.save(str(alt))
        log.warning(f"Fichier original verrouillé — sauvegardé sous : {alt}")
        fichier_sortie = alt
    return fichier_sortie


if __name__ == "__main__":
    fichier = remplir_excel()
    if fichier:
        print(f"\nFichier généré : {fichier}")
    else:
        print("\nAucun fichier généré.")
