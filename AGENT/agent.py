"""
SBEAE Agent - Suivi hebdomadaire techniciens terrain
=====================================================
Cet agent tourne en arrière-plan et gère :
- Rappels automatiques à 20h si saisie manquante
- Récap journalier envoyé aux chargés d'affaires à 20h30
- Compilation vendredi soir + remplissage Excel + envoi mails fournisseurs

Lancement : python agent.py
"""

import json
import os
import smtplib
import schedule
import time
import logging
import anthropic
import resend
from firebase_client import (
    charger_techniciens as fb_charger_techniciens,
    charger_charges_affaires as fb_charger_charges_affaires,
    charger_saisies_du_jour,
    charger_saisies_semaine,
)
from datetime import datetime, date
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

# ─── CONFIGURATION ────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR.parent / "data"

TEST_EMAIL_REDIRECT = None  # Si défini, tous les mails sont redirigés vers cette adresse
CONFIG_PATH = BASE_DIR / "config.json"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(BASE_DIR / "agent.log"),
        logging.StreamHandler()
    ]
)
log = logging.getLogger(__name__)

# ─── CHARGEMENT DES DONNÉES ───────────────────────────────────────────────────
def charger_config():
    env_config = os.environ.get("AGENT_CONFIG")
    if env_config:
        return json.loads(env_config)
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)

def charger_techniciens():
    return fb_charger_techniciens()

def charger_charges_affaires():
    return fb_charger_charges_affaires()

def charger_saisies():
    raise NotImplementedError("charger_saisies() supprimée — utiliser charger_saisies_du_jour() ou charger_saisies_semaine() depuis firebase_client.")

def sauvegarder_saisies(saisies):
    raise NotImplementedError("sauvegarder_saisies() supprimée — les saisies sont écrites par l'appli web dans Firestore.")

# ─── UTILITAIRES ──────────────────────────────────────────────────────────────
def get_date_aujourdhui():
    return date.today().isoformat()

def get_semaine_courante():
    return date.today().isocalendar()[1]

def get_annee_courante():
    return date.today().year

def est_vendredi():
    return date.today().weekday() == 4

def techniciens_non_saisis(techniciens, saisies):
    """Retourne la liste des techniciens qui n'ont pas saisi aujourd'hui."""
    aujourd_hui = get_date_aujourdhui()
    ids_saisis = {s["tech_id"] for s in saisies if s["date"] == aujourd_hui}
    return [t for t in techniciens if t["uid"] not in ids_saisis]

def saisies_du_jour(saisies):
    raise NotImplementedError("saisies_du_jour() supprimée — utiliser charger_saisies_du_jour() depuis firebase_client.")

def saisies_de_la_semaine(saisies):
    raise NotImplementedError("saisies_de_la_semaine() supprimée — utiliser charger_saisies_semaine() depuis firebase_client.")

# ─── ENVOI D'EMAILS ───────────────────────────────────────────────────────────
def envoyer_email(config, destinataire_email, destinataire_nom, sujet, corps_html, pieces_jointes=None):
    """Envoie un email via l'API Resend."""
    try:
        if TEST_EMAIL_REDIRECT:
            sujet = f"[TEST >> {destinataire_email}] {sujet}"
            destinataire_email = TEST_EMAIL_REDIRECT

        resend.api_key = os.environ.get("RESEND_API_KEY", "")

        params = {
            "from": "noreply@resend.dev",
            "to": [destinataire_email],
            "subject": sujet,
            "html": corps_html,
        }

        if pieces_jointes:
            attachments = []
            for chemin_fichier in pieces_jointes:
                with open(chemin_fichier, "rb") as f:
                    import base64
                    attachments.append({
                        "filename": Path(chemin_fichier).name,
                        "content": base64.b64encode(f.read()).decode("utf-8"),
                    })
            params["attachments"] = attachments

        resend.Emails.send(params)
        log.info(f"Email envoyé à {destinataire_nom} ({destinataire_email}) : {sujet}")
        return True

    except Exception as e:
        log.error(f"Erreur envoi email à {destinataire_email} : {e}")
        return False

# ─── SUPPRESSION COMPTES AUTH ─────────────────────────────────────────────────
def traiter_suppressions():
    """Supprime les comptes Firebase Auth en queue (écrits par l'appli web)."""
    from firebase_admin import auth as fb_auth
    from firebase_client import _get_db
    db = _get_db()
    docs = list(db.collection("suppressions").stream())
    if not docs:
        return
    for snap in docs:
        uid = snap.to_dict().get("uid") or snap.id
        try:
            fb_auth.delete_user(uid)
            log.info(f"Compte Auth supprimé : {uid}")
        except Exception as e:
            log.error(f"Erreur suppression Auth {uid} : {e}")
        snap.reference.delete()

# ─── CHAMPS PAR MISSION ───────────────────────────────────────────────────────
LABELS_CHAMPS = {
    "cpt_dn15_20":       "CPT DN15-20",
    "cpt_dn_sup20":      "CPT DN>20",
    "cpt_dn30_40":       "CPT DN30-40",
    "modules_poses":     "Modules",
    "modules_relances":  "Modules relancés",
    "rac":               "RAC",
    "mo_heures":         "MO (h)",
    "clapets_fact":      "Clapets fact.",
    "cpt_releves":       "CPT relevés",
    "infructueux":       "Infructueux",
    "absents_pda":       "Abs. PDA",
    "ctrl_vente_inf10":  "Ctrl AC <10pts",
    "ctrl_vente_sup10":  "Ctrl AC >10pts",
    "ctrl_contrat_inf10":"Ctrl Ctr <10pts",
    "ctrl_contrat_sup10":"Ctrl Ctr >10pts",
    "clients_absents":   "Cl. absents",
    "pi_visites":        "PI visités",
    "pi_conformes":      "PI conformes",
    "pi_non_conformes":  "PI non conf.",
    "pi_inaccessibles":  "PI inaccessibles",
    "anc_controles":     "ANC contrôlés",
    "anc_conformes":     "ANC conformes",
    "anc_non_conformes": "ANC non conf.",
    "total_heures":      "Heures",
}

CHAMPS_PAR_MISSION = {
    "RT_Compteur_Module": ["cpt_dn15_20", "cpt_dn_sup20", "modules_poses", "rac", "mo_heures"],
    "Releve_CPT":         ["cpt_releves", "infructueux", "absents_pda"],
    "Controle_AC":        ["ctrl_vente_inf10", "ctrl_vente_sup10", "ctrl_contrat_inf10", "clients_absents"],
    "RT_CPT_Arras":       ["cpt_dn15_20", "cpt_dn_sup20", "rac"],
    "RT_CPT_SEPIG":       ["cpt_dn30_40", "rac", "mo_heures"],
    "RT_CPT_Suez":        ["cpt_dn15_20", "cpt_dn_sup20", "modules_poses"],
    "RT_CPT":             ["cpt_dn15_20", "cpt_dn_sup20", "rac"],
    "PI_Poteau_Incendie": ["pi_visites", "pi_conformes", "pi_non_conformes"],
    "Controle_ANC":       ["anc_controles", "anc_conformes", "anc_non_conformes"],
}

# ─── TÂCHE 1 : RAPPEL 20H AUX TECHNICIENS ────────────────────────────────────
def tache_rappel_techniciens():
    """Envoie un rappel aux techniciens qui n'ont pas saisi leur journée."""
    log.info("=== TÂCHE : Rappels techniciens ===")

    config = charger_config()
    techniciens = charger_techniciens()
    saisies = charger_saisies_du_jour()

    non_saisis = techniciens_non_saisis(techniciens, saisies)

    if not non_saisis:
        log.info("Tous les techniciens ont saisi leur journée. Aucun rappel nécessaire.")
        return

    log.info(f"{len(non_saisis)} technicien(s) sans saisie : {[t['nom'] for t in non_saisis]}")

    for tech in non_saisis:
        sujet = f"[EAE Flow] Rappel saisie journalière — {datetime.now().strftime('%d/%m/%Y')}"
        corps = f"""
        <html><body style="font-family: Arial, sans-serif; color: #333;">
        <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
            <div style="background: #1A56DB; color: white; padding: 16px 24px; border-radius: 8px 8px 0 0;">
                <h2 style="margin: 0; font-size: 18px;">EAE Flow — Rappel saisie</h2>
            </div>
            <div style="background: #f9f9f9; padding: 24px; border: 1px solid #e0e0e0; border-radius: 0 0 8px 8px;">
                <p>Bonjour <strong>{tech['prenom']} {tech['nom']}</strong>,</p>
                <p>Vous n'avez pas encore saisi vos statistiques du jour
                (<strong>{datetime.now().strftime('%A %d %B %Y')}</strong>).</p>
                <p>Merci de vous connecter à EAE Flow pour enregistrer votre journée.</p>
                <br>
                <p style="color: #666; font-size: 13px;">
                    Ce message est automatique. Ne pas répondre à cet email.
                </p>
            </div>
        </div>
        </body></html>
        """
        envoyer_email(config, tech["email"], f"{tech['prenom']} {tech['nom']}", sujet, corps)

# ─── ANALYSE CLAUDE ───────────────────────────────────────────────────────────
def generer_analyse_claude(ca, mes_techs, saisies_jour, non_saisis_ids):
    """Génère une analyse narrative de la journée via Claude API."""
    try:
        lignes = []
        for tech in mes_techs:
            saisie = next((s for s in saisies_jour if s["tech_id"] == tech["uid"]), None)
            mission = tech["mission"].replace("_", " ")
            if saisie:
                champs = {
                    "Heures": saisie.get("total_heures"),
                    "CPT DN15-20": saisie.get("cpt_dn15_20"),
                    "CPT DN>20": saisie.get("cpt_dn_sup20"),
                    "Modules posés": saisie.get("modules_poses"),
                    "CPT relevés": saisie.get("cpt_releves"),
                    "Infructueux": saisie.get("infructueux"),
                    "Ctrl AC <10pts": saisie.get("ctrl_vente_inf10"),
                    "Ctrl AC >10pts": saisie.get("ctrl_vente_sup10"),
                    "Paniers midi": saisie.get("paniers_midi"),
                    "Paniers soir": saisie.get("paniers_soir"),
                }
                stats_str = ", ".join(f"{k}: {v}" for k, v in champs.items() if v is not None)
                lignes.append(f"- {tech['prenom']} {tech['nom']} ({mission}) : {stats_str or 'données saisies'}")
            else:
                lignes.append(f"- {tech['prenom']} {tech['nom']} ({mission}) : NON SAISI")

        donnees = "\n".join(lignes)
        date_str = datetime.now().strftime("%A %d %B %Y")

        prompt = f"""Tu es analyste pour SBEAE, une entreprise de travaux sur compteurs d'eau (remplacement, relevé, contrôle).
Rédige une analyse narrative courte (3 à 5 phrases) en français pour le chargé d'affaires {ca['prenom']} {ca['nom']}, basée sur les saisies du {date_str}.

Données des techniciens :
{donnees}

Sois direct et factuel. Mets en avant les performances notables et les manques (saisies manquantes ou volumes faibles). Ne génère pas de HTML, juste du texte pur."""

        client_claude = anthropic.Anthropic()
        response = client_claude.messages.create(
            model="claude-opus-4-8",
            max_tokens=512,
            messages=[{"role": "user", "content": prompt}]
        )
        texte = response.content[0].text.strip()
        log.info(f"Analyse Claude generee pour {ca['prenom']} {ca['nom']}")
        return texte

    except Exception as e:
        log.error(f"Erreur analyse Claude pour {ca.get('nom', '?')} : {e}")
        return ""


# ─── TÂCHE 2 : RÉCAP JOURNALIER AUX CHARGÉS D'AFFAIRES ──────────────────────
def tache_recap_journalier():
    """Envoie le récap du jour à chaque chargé d'affaires."""
    log.info("=== TÂCHE : Récap journalier chargés d'affaires ===")

    config = charger_config()
    techniciens = charger_techniciens()
    charges = charger_charges_affaires()
    saisies_jour = charger_saisies_du_jour()
    non_saisis = techniciens_non_saisis(techniciens, saisies_jour)
    non_saisis_ids = {t["uid"] for t in non_saisis}

    for ca in charges:
        mes_techs = [t for t in techniciens if t["charge_id"] == ca["id"]]
        if not mes_techs:
            continue

        # Construire le tableau HTML des techniciens
        lignes_html = ""
        for tech in mes_techs:
            saisie = next((s for s in saisies_jour if s["tech_id"] == tech["uid"]), None)
            statut_couleur = "#16A34A" if saisie else "#DC2626"
            statut_texte = "Saisi" if saisie else "Non saisi"

            if saisie:
                stats = _formater_stats_mission(saisie, tech.get("mission", ""))
            else:
                stats = "<em style='color:#999'>—</em>"

            lignes_html += f"""
            <tr style="border-bottom: 1px solid #f0f0f0; background: {'#fff' if saisie else '#fff5f5'};">
                <td style="padding: 10px 16px; font-weight: 600;">{tech['prenom']} {tech['nom']}</td>
                <td style="padding: 10px 16px; color: #666; font-size: 13px;">{tech['mission'].replace('_', ' ')}</td>
                <td style="padding: 10px 16px; font-size: 13px;">{stats}</td>
                <td style="padding: 10px 16px; text-align: center;">
                    <span style="background: {statut_couleur}20; color: {statut_couleur}; padding: 3px 10px; border-radius: 5px; font-size: 12px; font-weight: 600;">
                        {statut_texte}
                    </span>
                </td>
            </tr>
            """

        nb_saisis = len(mes_techs) - len([t for t in mes_techs if t["uid"] in non_saisis_ids])
        nb_total = len(mes_techs)

        analyse = generer_analyse_claude(ca, mes_techs, saisies_jour, non_saisis_ids)
        bloc_analyse = f"""
                <div style="background: #F0F7FF; border-left: 4px solid #1A56DB; margin: 0 24px 16px; padding: 14px 16px; border-radius: 0 6px 6px 0; font-size: 14px; line-height: 1.6; color: #1e3a5f;">
                    <strong style="display: block; margin-bottom: 6px; font-size: 11px; text-transform: uppercase; letter-spacing: 0.05em; color: #1A56DB;">Analyse du jour · Claude</strong>
                    {analyse}
                </div>""" if analyse else ""

        sujet = f"[EAE Flow] Récap journalier — {datetime.now().strftime('%d/%m/%Y')} — {nb_saisis}/{nb_total} saisies"
        corps = f"""
        <html><body style="font-family: Arial, sans-serif; color: #333;">
        <div style="max-width: 700px; margin: 0 auto; padding: 20px;">
            <div style="background: #0F172A; color: white; padding: 16px 24px; border-radius: 8px 8px 0 0;">
                <h2 style="margin: 0; font-size: 18px;">EAE Flow — Récap journalier</h2>
                <p style="margin: 4px 0 0; color: #94A3B8; font-size: 13px;">
                    {datetime.now().strftime('%A %d %B %Y')} · {nb_saisis}/{nb_total} techniciens ont saisi
                </p>
            </div>
            <div style="background: #fff; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 8px 8px; overflow: hidden;">
                <p style="padding: 16px 24px 0; margin: 0;">Bonjour <strong>{ca['prenom']} {ca['nom']}</strong>,</p>
                <p style="padding: 8px 24px 16px; margin: 0; color: #666;">
                    Voici le bilan de la journée pour vos {nb_total} techniciens.
                </p>
                {bloc_analyse}
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <thead>
                        <tr style="background: #F8FAFC; border-bottom: 2px solid #e0e0e0;">
                            <th style="padding: 10px 16px; text-align: left; font-size: 11px; color: #999; text-transform: uppercase; letter-spacing: 0.05em;">Technicien</th>
                            <th style="padding: 10px 16px; text-align: left; font-size: 11px; color: #999; text-transform: uppercase; letter-spacing: 0.05em;">Mission</th>
                            <th style="padding: 10px 16px; text-align: left; font-size: 11px; color: #999; text-transform: uppercase; letter-spacing: 0.05em;">Stats du jour</th>
                            <th style="padding: 10px 16px; text-align: center; font-size: 11px; color: #999; text-transform: uppercase; letter-spacing: 0.05em;">Statut</th>
                        </tr>
                    </thead>
                    <tbody>
                        {lignes_html}
                    </tbody>
                </table>
                <p style="padding: 16px 24px; color: #999; font-size: 12px; border-top: 1px solid #f0f0f0;">
                    Ce message est envoyé automatiquement chaque soir à 20h30. Ne pas répondre.
                </p>
            </div>
        </div>
        </body></html>
        """
        envoyer_email(config, ca["email"], f"{ca['prenom']} {ca['nom']}", sujet, corps)

def _formater_stats_mission(saisie, mission):
    """Formate les stats d'une saisie selon la mission du technicien."""
    parts = []
    for key in CHAMPS_PAR_MISSION.get(mission, []):
        val = saisie.get(key)
        if val is not None:
            parts.append(f"{LABELS_CHAMPS.get(key, key)} : <strong>{val}</strong>")
    heures = saisie.get("total_heures")
    if heures is not None:
        parts.append(f"Heures : <strong>{heures}</strong>")
    html = " · ".join(parts) if parts else "Données saisies"
    commentaires = (saisie.get("commentaires") or "").strip()
    if commentaires:
        html += f'<br><em style="color:#666; font-size:12px;">💬 {commentaires}</em>'
    return html

# ─── TÂCHE 3 : COMPILATION VENDREDI ──────────────────────────────────────────
def tache_compilation_vendredi():
    """Compile les totaux de la semaine, remplit l'Excel, envoie les mails."""
    if not est_vendredi():
        log.info("Pas vendredi — compilation ignorée.")
        return

    log.info("=== TÂCHE : Compilation vendredi ===")

    config = charger_config()
    techniciens = charger_techniciens()
    charges = charger_charges_affaires()
    saisies_sem = charger_saisies_semaine()

    # Calculer les totaux par technicien (format agent.py)
    totaux = _calculer_totaux_semaine(techniciens, saisies_sem)

    # Remplir l'Excel via Claude (remplir_excel.py)
    fichier_excel = None
    try:
        from remplir_excel import remplir_excel
        fichier_excel = remplir_excel()
    except Exception as e:
        log.error(f"Erreur remplissage Excel : {e}")

    # Envoyer aux fournisseurs groupés par CA
    _envoyer_mails_fournisseurs(config, techniciens, charges, totaux)

    # Envoyer au patron avec l'Excel
    _envoyer_mail_patron(config, totaux, techniciens, fichier_excel)

    log.info("=== Compilation vendredi terminée ===")

def _calculer_totaux_semaine(techniciens, saisies_sem):
    """Calcule les totaux hebdomadaires pour chaque technicien."""
    totaux = {}
    champs_numeriques = [
        "total_heures", "paniers_midi", "paniers_soir", "rdv_jour",
        "cpt_dn15_20", "cpt_dn_sup20", "rac", "clapets_fact", "clapets_non_fact",
        "sr_laiton", "reducteurs", "mo_heures", "modules_poses", "modules_relances",
        "cpt_non_faits", "depl_injustifies", "clients_absents",
        "cpt_releves", "infructueux", "absents_pda",
        "ctrl_vente_inf10", "ctrl_vente_sup10", "ctrl_contrat_inf10", "ctrl_contrat_sup10",
    ]

    for tech in techniciens:
        mes_saisies = [s for s in saisies_sem if s["tech_id"] == tech["uid"]]
        total = {"tech_id": tech["uid"], "nb_jours": len(mes_saisies)}
        for champ in champs_numeriques:
            total[champ] = sum(s.get(champ, 0) or 0 for s in mes_saisies)
        totaux[tech["uid"]] = total

    return totaux

def _remplir_excel(config, techniciens, totaux):
    """Remplit le fichier Excel EAE avec les totaux de la semaine."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        fichier_source = BASE_DIR / config["fichiers"]["excel_suivi"]
        if not fichier_source.exists():
            log.warning(f"Fichier Excel source introuvable : {fichier_source}")
            return None

        wb = openpyxl.load_workbook(fichier_source)
        semaine = get_semaine_courante()
        annee = get_annee_courante()

        # Chercher l'onglet de la semaine courante
        nom_onglet = f"Sem_{semaine}"
        if nom_onglet not in wb.sheetnames:
            # Essayer de trouver un onglet similaire ou créer
            log.info(f"Onglet {nom_onglet} non trouvé, recherche d'un onglet existant...")
            # Trouver le dernier onglet Sem_X et en créer un nouveau
            onglets_sem = [s for s in wb.sheetnames if s.startswith("Sem_")]
            if onglets_sem:
                dernier = wb[onglets_sem[-1]]
                ws = wb.copy_worksheet(dernier)
                ws.title = nom_onglet
            else:
                ws = wb.create_sheet(nom_onglet)
        else:
            ws = wb[nom_onglet]

        # Trouver les lignes techniciens et remplir les totaux
        for tech in techniciens:
            total = totaux.get(tech["uid"], {})
            _remplir_ligne_technicien(ws, tech, total)

        # Sauvegarder
        dossier_sortie = BASE_DIR / config["fichiers"]["dossier_sortie"]
        dossier_sortie.mkdir(exist_ok=True)
        fichier_sortie = dossier_sortie / f"SUIVI_S{semaine}_{annee}.xlsx"
        wb.save(fichier_sortie)
        log.info(f"Excel sauvegardé : {fichier_sortie}")
        return fichier_sortie

    except ImportError:
        log.error("openpyxl non installé. Lancer : pip install openpyxl")
        return None
    except Exception as e:
        log.error(f"Erreur remplissage Excel : {e}")
        return None

def _remplir_ligne_technicien(ws, tech, total):
    """Trouve la ligne du technicien dans la feuille et remplit ses totaux."""
    nom_complet = f"{tech['nom']} {tech['prenom']}"
    nom_inverse = f"{tech['prenom']} {tech['nom']}"

    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value or "").strip().upper()
            if tech["nom"].upper() in val and (tech["prenom"].upper() in val or tech["prenom"][0].upper() in val):
                # Trouvé la ligne — remplir les colonnes adjacentes
                row_num = cell.row
                col_num = cell.column

                # Remplir les totaux selon la mission
                if tech["mission"] == "RT_Compteur_Module":
                    _remplir_rt_compteur(ws, row_num, col_num, total)
                elif tech["mission"] == "Releve_CPT":
                    _remplir_releve_cpt(ws, row_num, col_num, total)
                elif tech["mission"] == "Controle_AC":
                    _remplir_controle_ac(ws, row_num, col_num, total)

                log.info(f"Technicien {tech['nom']} trouvé à la ligne {row_num}")
                return

    log.warning(f"Technicien {tech['nom']} non trouvé dans le fichier Excel")

def _remplir_rt_compteur(ws, row, col_debut, total):
    """Remplit les colonnes RT Compteur pour une ligne technicien."""
    # Mapping colonnes — à ajuster selon la structure réelle du fichier
    mapping = {
        1: total.get("cpt_dn15_20", 0),
        2: total.get("cpt_dn_sup20", 0),
        3: total.get("modules_poses", 0),
        4: total.get("modules_relances", 0),
        5: total.get("rac", 0),
        6: total.get("clapets_fact", 0),
        7: total.get("mo_heures", 0),
        8: total.get("total_heures", 0),
    }
    for offset, valeur in mapping.items():
        ws.cell(row=row, column=col_debut + offset, value=valeur)

def _remplir_releve_cpt(ws, row, col_debut, total):
    """Remplit les colonnes Relevé CPT."""
    mapping = {
        1: total.get("cpt_releves", 0),
        2: total.get("infructueux", 0),
        3: total.get("absents_pda", 0),
        4: total.get("total_heures", 0),
    }
    for offset, valeur in mapping.items():
        ws.cell(row=row, column=col_debut + offset, value=valeur)

def _remplir_controle_ac(ws, row, col_debut, total):
    """Remplit les colonnes Contrôle AC."""
    mapping = {
        1: total.get("ctrl_vente_inf10", 0),
        2: total.get("ctrl_vente_sup10", 0),
        3: total.get("ctrl_contrat_inf10", 0),
        4: total.get("ctrl_contrat_sup10", 0),
        5: total.get("total_heures", 0),
    }
    for offset, valeur in mapping.items():
        ws.cell(row=row, column=col_debut + offset, value=valeur)

def _envoyer_mails_fournisseurs(config, techniciens, charges, totaux):
    """Envoie les récaps hebdo aux chargés d'affaires."""
    semaine = get_semaine_courante()

    for ca in charges:
        mes_techs = [t for t in techniciens if t["charge_id"] == ca["id"]]
        if not mes_techs:
            continue

        lignes_html = ""
        for tech in mes_techs:
            total = totaux.get(tech["uid"], {})
            nb_jours = total.get("nb_jours", 0)

            lignes_html += f"""
            <tr style="border-bottom: 1px solid #f0f0f0;">
                <td style="padding: 10px 16px; font-weight: 600;">{tech['prenom']} {tech['nom']}</td>
                <td style="padding: 10px 16px; color: #666; font-size: 13px;">{tech['mission'].replace('_', ' ')}</td>
                <td style="padding: 10px 16px; text-align: center; font-size: 13px;">{nb_jours}/5</td>
                <td style="padding: 10px 16px; text-align: center; font-weight: 700; color: #1A56DB;">
                    {total.get('cpt_dn15_20', 0) or total.get('cpt_releves', 0) or total.get('ctrl_vente_inf10', 0)}
                </td>
                <td style="padding: 10px 16px; text-align: center;">{total.get('modules_poses', 0) or '—'}</td>
                <td style="padding: 10px 16px; text-align: center;">{total.get('total_heures', 0)}h</td>
            </tr>
            """

        sujet = f"[EAE Flow] Bilan semaine S{semaine} — {len(mes_techs)} techniciens"
        corps = f"""
        <html><body style="font-family: Arial, sans-serif; color: #333;">
        <div style="max-width: 700px; margin: 0 auto; padding: 20px;">
            <div style="background: #0F172A; color: white; padding: 16px 24px; border-radius: 8px 8px 0 0;">
                <h2 style="margin: 0; font-size: 18px;">EAE Flow — Bilan semaine S{semaine}</h2>
                <p style="margin: 4px 0 0; color: #94A3B8; font-size: 13px;">
                    Totaux hebdomadaires · {len(mes_techs)} techniciens
                </p>
            </div>
            <div style="background: #fff; border: 1px solid #e0e0e0; border-top: none; border-radius: 0 0 8px 8px; overflow: hidden;">
                <p style="padding: 16px 24px 0; margin: 0;">Bonjour <strong>{ca['prenom']} {ca['nom']}</strong>,</p>
                <p style="padding: 8px 24px 16px; margin: 0; color: #666;">
                    Voici les totaux de la semaine S{semaine} pour vos techniciens.
                </p>
                <table style="width: 100%; border-collapse: collapse; font-size: 14px;">
                    <thead>
                        <tr style="background: #F8FAFC; border-bottom: 2px solid #e0e0e0;">
                            <th style="padding: 10px 16px; text-align: left; font-size: 11px; color: #999; text-transform: uppercase;">Technicien</th>
                            <th style="padding: 10px 16px; text-align: left; font-size: 11px; color: #999; text-transform: uppercase;">Mission</th>
                            <th style="padding: 10px 16px; text-align: center; font-size: 11px; color: #999; text-transform: uppercase;">Jours</th>
                            <th style="padding: 10px 16px; text-align: center; font-size: 11px; color: #999; text-transform: uppercase;">CPT / Unités</th>
                            <th style="padding: 10px 16px; text-align: center; font-size: 11px; color: #999; text-transform: uppercase;">Modules</th>
                            <th style="padding: 10px 16px; text-align: center; font-size: 11px; color: #999; text-transform: uppercase;">Heures</th>
                        </tr>
                    </thead>
                    <tbody>{lignes_html}</tbody>
                </table>
                <p style="padding: 16px 24px; color: #999; font-size: 12px; border-top: 1px solid #f0f0f0;">
                    Message automatique envoyé chaque vendredi soir. Le fichier Excel complet est transmis au responsable.
                </p>
            </div>
        </div>
        </body></html>
        """
        envoyer_email(config, ca["email"], f"{ca['prenom']} {ca['nom']}", sujet, corps)

def _envoyer_mail_patron(config, totaux, techniciens, fichier_excel):
    """Envoie l'Excel compilé au patron."""
    semaine = get_semaine_courante()
    patron = config["patron"]

    sujet = f"[EAE Flow] Suivi hebdomadaire S{semaine} — Excel joint"
    corps = f"""
    <html><body style="font-family: Arial, sans-serif; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: #0F172A; color: white; padding: 16px 24px; border-radius: 8px 8px 0 0;">
            <h2 style="margin: 0; font-size: 18px;">EAE Flow — Suivi S{semaine}</h2>
        </div>
        <div style="background: #fff; border: 1px solid #e0e0e0; border-top: none; padding: 24px; border-radius: 0 0 8px 8px;">
            <p>Bonjour <strong>{patron['nom']}</strong>,</p>
            <p>Veuillez trouver ci-joint le fichier Excel de suivi hebdomadaire pour la semaine S{semaine}.</p>
            <p>
                <strong>{len(techniciens)}</strong> techniciens · 
                <strong>{sum(1 for t in techniciens if totaux.get(t['id'], {}).get('nb_jours', 0) > 0)}</strong> ont saisi au moins une journée
            </p>
            <p style="color: #999; font-size: 12px; margin-top: 24px;">Message automatique EAE Flow.</p>
        </div>
    </div>
    </body></html>
    """

    pieces_jointes = [str(fichier_excel)] if fichier_excel else None
    envoyer_email(config, patron["email"], patron["nom"], sujet, corps, pieces_jointes)

# ─── PLANIFICATEUR ────────────────────────────────────────────────────────────
def demarrer_agent():
    """Démarre l'agent et planifie les tâches."""
    config = charger_config()
    horaires = config["horaires"]

    log.info("=" * 50)
    log.info("EAE Flow Agent démarré")
    log.info(f"Rappels techniciens    : {horaires['rappel_saisie']}")
    log.info(f"Récap chargés d'aff.  : {horaires['recap_journalier']}")
    log.info(f"Compilation vendredi   : {horaires['compilation_vendredi']}")
    log.info("=" * 50)

    # Traiter les suppressions de comptes en attente
    traiter_suppressions()

    # Planifier les tâches
    schedule.every().day.at(horaires["rappel_saisie"]).do(tache_rappel_techniciens)
    schedule.every().day.at(horaires["recap_journalier"]).do(tache_recap_journalier)
    schedule.every().friday.at(horaires["compilation_vendredi"]).do(tache_compilation_vendredi)

    # Boucle principale
    while True:
        schedule.run_pending()
        time.sleep(30)

# ─── MODE TEST ────────────────────────────────────────────────────────────────
def tester_agent(redirect_email=None):
    """Lance toutes les tâches immédiatement pour tester."""
    global TEST_EMAIL_REDIRECT
    TEST_EMAIL_REDIRECT = redirect_email
    if redirect_email:
        log.info(f"=== MODE TEST — Tous les mails redirigés vers {redirect_email} ===")
    else:
        log.info("=== MODE TEST — Exécution immédiate de toutes les tâches ===")
    tache_rappel_techniciens()
    tache_recap_journalier()
    tache_compilation_vendredi()
    log.info("=== TEST TERMINÉ ===")

# ─── POINT D'ENTRÉE ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == "test":
        redirect = sys.argv[2] if len(sys.argv) > 2 else "marin.balcon.sbeae@gmail.com"
        tester_agent(redirect_email=redirect)
    else:
        demarrer_agent()
